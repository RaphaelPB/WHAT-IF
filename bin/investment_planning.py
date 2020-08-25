# -*- coding: utf-8 -*-
"""
Created on Fri Feb 09 13:43:36 2018

@author: rapy
"""
from __future__ import division #As in python 2.7 division between integers gives an integer - avoid problems
from result_analysis import ResultAnalysis #Exports results to excel sheets
import pandas as pd
from openpyxl import load_workbook

class InvestmentPlan():
    def __init__(self,model):
        self.CostBenefitAnalysis={'Single Investments':{},'Total Plan':{},'In Plan':{}}
        self.memINVEST={}
        self.model=model
#-------Fix and remember selected investment to switch to fully linear model
        for ip in self.model.ninvphase:
            for inv in self.model.ninvest:   
                self.model.IbINVEST[ip,inv].fixed = True
                self.memINVEST[ip,inv]=self.model.IbINVEST[ip,inv].value

#-------Export results to excel files
    def exportresults(self,parameters,solver,outpath):            
        #Solve model with fixed investments
        solverstatus=solver.solve(self.model)
        #Export results with shadowprices
        self.results=ResultAnalysis()
        self.results.readresults(self.model,parameters,solverstatus)
        self.results.exportresults(outpath,parameters)    
        
        #Save investment plan for export
        self.InvestPlan=self.results.InvestmentPlan_inv_ip
        self.Investments=self.results.Investments_ip
        print('---------------INVESTMENTS-------------------')
        
#-------CBA of selected investments at their implementation time--------
#    def cba_investinplan(self,parameters,solver):
        print('CBA of selected investments at their implementation time')
        self.InvestmentAnalysis1={}
        self.InvestmentAnalysis1['optimal']=self.results.EconomicBalance['Balance [$]']
        for inv in self.model.ninvest:
            for ip in self.model.ninvphase:
                if self.memINVEST[ip,inv]==1:
                    self.model.IbINVEST[ip,inv] = 0
                    self.model.IbINVEST[ip,inv].fixed = True
                    solverstatus=solver.solve(self.model)
                    self.results.readresults(self.model,parameters,solverstatus)
                    self.InvestmentAnalysis1[inv]=self.results.EconomicBalance['Balance [$]']
                    benefit=(self.InvestmentAnalysis1['optimal']-self.InvestmentAnalysis1[inv])*len(self.model.nyear)
                    cost=self.model.iInvCost[inv]
                    self.CostBenefitAnalysis['In Plan'][inv]=benefit-cost
                    print('Investment ' + inv + ' , net benefit [$]:')
                    print(benefit-cost)
                    self.model.IbINVEST[ip,inv] = 1
                    self.model.IbINVEST[ip,inv].fixed = True
        
#-------CBA of single investments at first investment phase-----------
#    def cba_investsingle(self,parameters,solver):
        print('Single investment at first investment phase')    
        #reset investments
        for inv in self.model.ninvest: 
            for ip in self.model.ninvphase:
                self.model.IbINVEST[ip,inv] = 0
                self.model.IbINVEST[ip,inv].fixed = True
        #solve buisness as usual case (no investments)
        self.InvestmentAnalysis2={}
        solverstatus=solver.solve(self.model)
        self.results.readresults(self.model,parameters,solverstatus) 
        self.InvestmentAnalysis2['base']=self.results.EconomicBalance['Balance [$]']

        #analyse single investments
        for inv in self.model.ninvest:
            self.model.IbINVEST['ip1',inv] = 1
            self.model.IbINVEST['ip1',inv].fixed = True
            solverstatus=solver.solve(self.model)
            self.results.readresults(self.model,parameters,solverstatus)
            self.InvestmentAnalysis2[inv]=self.results.EconomicBalance['Balance [$]']
            benefit=(self.InvestmentAnalysis2[inv]-self.InvestmentAnalysis2['base'])*len(self.model.nyear)
            cost=self.model.iInvCost[inv]
            self.CostBenefitAnalysis['Single Investments'][inv]=benefit-cost
            print('Investment ' + inv + ' net benefit [$]:')
            print(benefit-cost)
            self.model.IbINVEST['ip1',inv] = 0
            self.model.IbINVEST['ip1',inv].fixed = True            
            
#-------CBA of entire investment plan--------
        benefit = (self.InvestmentAnalysis1['optimal']-self.InvestmentAnalysis2['base'])*len(self.model.nyear)
        cost = sum(self.model.iInvCost[inv] for inv in self.model.ninvest if sum(self.memINVEST[ip,inv] for ip in self.model.ninvphase) == 1)
        self.CostBenefitAnalysis['Total Plan'][0]=benefit-cost       
        
#-------Export results----------------------
        book = load_workbook(outpath)
        writer = pd.ExcelWriter(outpath, engine='openpyxl') #able to save formulas (graphs die however)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)        
        
        data=[self.InvestPlan, self.Investments, self.CostBenefitAnalysis['Total Plan'], self.CostBenefitAnalysis['In Plan'], self.CostBenefitAnalysis['Single Investments']]
        title=['Invest sequence','Investments [$]','CBA of investment plan [$]','CBA of investments in plan[$]','CBA of single investment options [$]']
        self.results.exportsheet(writer, 'Investments', data, title, index=[0,1,2,3,4])
        writer.save()
        
#------Differenciated results---------------
    def singleinvestmentresults(self,parameters,investment,solver,outpath):
        #Solve with optimal investment plan
        for inv in self.model.ninvest:
            for ip in self.model.ninvphase:
                if self.memINVEST[ip,inv]==1:
                    self.model.IbINVEST[ip,inv] = 1
                    self.model.IbINVEST[ip,inv].fixed = True
                else:
                    self.model.IbINVEST[ip,inv] = 0
                    self.model.IbINVEST[ip,inv].fixed = True
        #Save results
        solverstatus=solver.solve(self.model)
        Optimal=ResultAnalysis()
        Optimal.readresults(self.model,parameters,solverstatus) 
        
        #Solve without investment
        for ip in self.model.ninvphase:
            if self.memINVEST[ip,investment]==1:
                self.model.IbINVEST[ip,investment] = 0
                self.model.IbINVEST[ip,investment].fixed = True
        #Save results
        solverstatus=solver.solve(self.model)
        NoInvest=ResultAnalysis()
        NoInvest.readresults(self.model,parameters,solverstatus)
     
        #Export results
        months=['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']        
        farmingzone=['Crop Production [t/y]','Benefits [$/y]','Cultivation costs [$/y]','Irrigation costs [$/y]','Transport cost [$/y]','Cultivated land [ha]','Irrig net abstraction [m3/y]','Irrig withdrawals [m3/y]','Irrigation losses [m3/y]','Irrigation return flow [m3/y]','Irrig. dem. satisfaction [%]','Rainfed dem. satisfaction [%]','Land shadowprice [$/ha]','Irrig. land shadowprice [$/ha]','Rain. land shadowprice [$/ha]']
        book = load_workbook(outpath)
        writer = pd.ExcelWriter(outpath, engine='openpyxl') #able to save formulas (graphs die however)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        #INVESTMENTS
        data=[Optimal.InvestmentPlan_inv_ip,Optimal.Investments_ip]
        dataref=[NoInvest.InvestmentPlan_inv_ip,NoInvest.Investments_ip]
        title=['Invest sequence','Investments [$]']
        Optimal.exportsheet(writer,'Investments',data,title,dataref=dataref,index=[0,1])
        
        #POWER NETWORK
        data=[Optimal.TmsLineTable]
        dataref=[NoInvest.TmsLineTable]
        title=['Transmission lines']
        Optimal.exportsheet(writer,'PowerNetwork',data,title,dataref=dataref)
            
        #POWER PLANTS
        data=[Optimal.PowerplantTable,Optimal.PpProd_pp_m,Optimal.PpProd_pp_y,Optimal.PpBenefits_pp_y]
        dataref=[Optimal.PowerplantTable,Optimal.PpProd_pp_m,Optimal.PpProd_pp_y,Optimal.PpBenefits_pp_y]
        title=['Power plant','Power production [kWh/month]','Power production [kWh/y]','Net Benefits* [$/y]']
        Optimal.exportsheet(writer,'PowerPlants',data,title,dataref=dataref,order={1:months})
            
        #ENERGY
        data=[Optimal.EnergyBalance,Optimal.EnergyBalance_co,Optimal.EnergyBalance_ca,Optimal.EnergyBalance_m,Optimal.EnergyBalance_y]
        dataref=[NoInvest.EnergyBalance,NoInvest.EnergyBalance_co,NoInvest.EnergyBalance_ca,NoInvest.EnergyBalance_m,NoInvest.EnergyBalance_y]
        title=['ENERGY BALANCE','ENERGY BALANCE','ENERGY BALANCE','ENERGY BALANCE','ENERGY BALANCE']
        Optimal.exportsheet(writer,'Energy',data,title,dataref=dataref,index=[0,1,2,3,4,5],order={3:months},total={3:1})
        
        #CROPS
        data=[Optimal.CropPrice_cr_y,Optimal.CropPrice_cr_cm,Optimal.CropProduction_cr_co,Optimal.LandUse_cr_co,Optimal.LandUse_cr_y,Optimal.LandUse_co_ft,Optimal.LandUse_cr_ft]
        dataref=[NoInvest.CropPrice_cr_y,NoInvest.CropPrice_cr_cm,NoInvest.CropProduction_cr_co,NoInvest.LandUse_cr_co,NoInvest.LandUse_cr_y,NoInvest.LandUse_co_ft,NoInvest.LandUse_cr_ft]
        title=['Crop Price [$/t]','Crop Price [$/t]','Crop Production [t]','Land Use [ha]','Land Use [ha]','Land Use [ha]','Land Use [ha]']
        Optimal.exportsheet(writer,'Crops',data,title,dataref=dataref)
        
        #FARMING ZONES
        data=[Optimal.FarmZonesTable,Optimal.CropProduction_fz_cr,Optimal.CropBenefits_fz_ft,Optimal.CropBenefits_fz_y,Optimal.CropTransport_fz_cm, Optimal.CropLoss_fz_cm]
        dataref=[NoInvest.FarmZonesTable,NoInvest.CropProduction_fz_cr,NoInvest.CropBenefits_fz_ft,NoInvest.CropBenefits_fz_y,NoInvest.CropTransport_fz_cm, NoInvest.CropLoss_fz_cm]
        title=['Farming zone','Crop production [t/y]','Net benefits [$/y]','Net benefits [$/y]','Crop Transport [t/y]', 'Crop Transport losses[t/y]']
        Optimal.exportsheet(writer,'FarmingZones',data,title,dataref=dataref,order={0:farmingzone})        
        
        #RESERVOIRS
        data=[Optimal.ReservoirTable,Optimal.ResStor_res_m]
        dataref=[NoInvest.ReservoirTable,NoInvest.ResStor_res_m]
        title=['Reservoir','Average storage']
        Optimal.exportsheet(writer,'Reservoirs',data,title,dataref=dataref,order={1:months})
        
        #HYDROPOWER
        data=[Optimal.HydropowerTable,Optimal.HpProd_hp_m,Optimal.HpStor_hp_m,Optimal.HpDis_hp_m,Optimal.HpProd_hp_y,Optimal.HpBenefits_hp_y]
        dataref=[NoInvest.HydropowerTable,NoInvest.HpProd_hp_m,NoInvest.HpStor_hp_m,NoInvest.HpDis_hp_m,NoInvest.HpProd_hp_y,NoInvest.HpBenefits_hp_y]
        title=['Hydropower','Power production [kWh]','Storage* [m3]','Discharge [m3]','Power production [kWh/y]', 'Net Benefits* [$/year]']
        Optimal.exportsheet(writer,'Hydropower',data,title,dataref=dataref,order={1:months,2:months,3:months},total={1:2,3:2,4:2,5:2})
        
        #MASS BALANCES
        data=[Optimal.EconomicBalance,Optimal.EconomicBalance_co,Optimal.EconomicBalance_ca,
              Optimal.WaterBalance,Optimal.WaterBalance_co,Optimal.WaterBalance_ca,
              Optimal.EnergyBalance,Optimal.EnergyBalance_co,Optimal.EnergyBalance_ca,
              Optimal.CropBalance,Optimal.CropBalance_co,Optimal.CropBalance_ca]
        dataref=[NoInvest.EconomicBalance,NoInvest.EconomicBalance_co,NoInvest.EconomicBalance_ca,
              NoInvest.WaterBalance,NoInvest.WaterBalance_co,NoInvest.WaterBalance_ca,
              NoInvest.EnergyBalance,NoInvest.EnergyBalance_co,NoInvest.EnergyBalance_ca,
              NoInvest.CropBalance,NoInvest.CropBalance_co,NoInvest.CropBalance_ca]
        title=['ECONOMIC BALANCE','ECONOMIC BALANCE','ECONOMIC BALANCE',
               'WATER BALANCE','WATER BALANCE','WATER BALANCE',
               'ENERGY BALANCE','ENERGY BALANCE','ENERGY BALANCE',
               'CROP BALANCE','CROP BALANCE','CROP BALANCE']
        Optimal.exportsheet(writer,'MassBalance',data,title,dataref=dataref,index=[0,1,2,3,4,5,6,7,8,9,10,11],cols=3)
        
        #MAIN
        Optimal.exportsheet(writer,'Main',[Optimal.MainTable],['Configurations'],index=[0])
        
        #Save excel files
        writer.save()